import openpyxl
import argparse
import asyncio
import aiofiles

async def extract_ip_addresses(filename):
    ip_address_ports = {}

    async with aiofiles.open(filename, 'r') as file:
        ip = ""
        async for line in file:
            if line.startswith("Nmap scan report for ") and "[host down, received no-response]" not in line and "[host, down, received host-unreach]" not in line:
                ip = line.split("Nmap scan report for ")[-1].replace('(', ',').replace(')','').strip()
            elif "open" in line and "Discovered" not in line and "tcp" in line:
                if ip in ip_address_ports:
                    ip_address_ports[ip].append(line.strip())
                else:
                    ip_address_ports[ip] = [line.strip()]

    return ip_address_ports

 
def write_to_excel(filename, ip_address_ports):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Host"
    ws['B1'] = "Port, service"

    row = 2
    for ip, ports in ip_address_ports.items():
        ws.cell(row=row, column=1, value=ip)
        ws.cell(row=row, column=2, value='\n'.join(ports).replace("syn-ack", '').replace('tcp', 'TCP').replace('?', ''))
        row += 1

    wb.save(filename)

 
async def main(input_filenames, output_filename):
     # Сбор результатов из всех файлов
    results = {}
    for filename in input_filenames:
        ip_address_ports = await extract_ip_addresses(filename)
        results.update(ip_address_ports)


    write_to_excel(output_filename, results)

    print("Успешно сохранено в файл ", output_filename)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Process multiple nmap files asynchronously and output results to an Excel file.')

    parser.add_argument('input_filenames', nargs='+', help='Input file(s) containing nmap scan results')
    parser.add_argument('output_filename', help='Output Excel file to write the results to')
    args = parser.parse_args()

    asyncio.run(main(args.input_filenames, args.output_filename))
